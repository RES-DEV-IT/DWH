from datetime import datetime, timedelta
from airflow.decorators import task
from airflow.providers.postgres.hooks.postgres import PostgresHook
from airflow.operators.bash import BashOperator
from airflow import DAG
from pyairtable import Table, Api
import re
import pandas as pd
import requests
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from send_file_by_mail import create_service, send_email_with_excel


CRON_EXP = None
START_DATE = datetime(2025, 11, 27, 4, 0, 0, 0)

default_args = {
    "owner": "Artem",
    "retries": 0
}

# API TOKENS
API_TOKENS = {
    "Portal": "patpHD9iVIlsFGE2w.80f76907cb8bdc1dffd465c3eb3f275bc26a7e247727bdd5559b07decc0eb7d9"
}

def get_table(
        token_name: str,
        base_name: str,
        table_name: str
) -> Table:
    """
    Функция для получения AirTable таблицы по её base_name и имени
    """

    # AIRTABLE API
    api_key = API_TOKENS[token_name]
    api = Api(api_key)
    available_bases = api.bases()

    for available_base in available_bases:
        if available_base.name == base_name:
            at_table = api.table(available_base.id, table_name)

            return at_table

    raise ValueError(f"Can't find base with name: {base_name}")

class KKSTransformer():
    def __init__(self) -> None:
        self.pattern_0 = r'^[a-zA-Z0-9]+$'
        self.pattern_1 = r'\d+\(\d+(?:\s*,\s*\d+)*\)[a-zA-Z0-9]*'
        self.pattern_2 = r'\d+[a-zA-Z0-9]*\(\d+(?:\s*,\s*\d+)*\)'
        self.pattern_3 = r'\d+\(\d+(?:\s*,\s*\d+)*\)[a-zA-Z0-9]*\(\d+(?:\s*,\s*\d+)*\)'

    def __call__(self, column_as_list):
        print(column_as_list)
        transformed_list = []
        for value in column_as_list:
            if type(value) is str:
                raw_kks_list = re.split(r'[\n\s]+', value.strip())

                kks = []
                for raw_kks in raw_kks_list:
                    kks.extend(self.process_one_kks(raw_kks))
                transformed_list.append("\n".join(kks))
            else:
                transformed_list.append(None) # NULL in postgres

        return transformed_list

    def process_one_kks(self, kks: str):
        # 10(20)KBF71AA201(100,200,300)
        if re.match(self.pattern_3, kks) is not None:
            base, first_list, body, tail, second_list = self.parse_pattern_3(kks)
            return self.create_kks(base, first_list, body, tail, second_list)
        # 10(20)KBF71AA201
        elif re.match(self.pattern_1, kks) is not None:
            base, first_list, body, tail, second_list = self.parse_pattern_1(kks)
            return self.create_kks(base, first_list, body, tail, second_list)
        # 10KBF71AA201(100,200,300)
        elif re.match(self.pattern_2, kks) is not None:
            base, first_list, body, tail, second_list = self.parse_pattern_2(kks)
            return self.create_kks(base, first_list, body, tail, second_list)
        # 10KBF71AA201
        elif re.match(self.pattern_0, kks) is not None:
            return [kks]

        # Other variants
        if kks in ["prototype", "(prototype)", "NA", "-"]:
            return [kks]
        assert 1==0, f"DON'T match patterns {kks}"

    def parse_pattern_1(self, kks):
        r = re.search(r'\(\d+(?:\s*,\s*\d+)*\)', kks)
        s, e = r.start(), r.end()
        kks = kks[:e+10] # 10(20)PCB03AA501KA02 изюавляемся от концовки -> 10(20)PCB03AA501
        base = kks[:s]
        first_list = kks[s+1:e-1].split(',')
        body = kks[e:-3]
        tail = kks[-3:]
        second_list = []
        return base, first_list, body, tail, second_list

    def parse_pattern_2(self, kks):
        r = re.search(r'\(\d+(?:\s*,\s*\d+)*\)', kks)
        s, e = r.start(), r.end()
        base = kks[:2]
        first_list = []
        body = kks[2:s-3]
        tail = kks[s-3:s]
        second_list = kks[s+1:e-1].split(',')
        return base, first_list, body, tail, second_list

    def parse_pattern_3(self, kks):
        s1, s2, e1, e2 = None, None, None, None
        for i, r in enumerate(re.finditer(r'\(\d+(?:\s*,\s*\d+)*\)', kks)):
            if i == 0: s1, e1 = (r.start(), r.end())
            if i == 1: s2, e2 = (r.start(), r.end())

        base = kks[:s1]
        first_list = kks[s1+1:e1-1].split(',')
        body = kks[e1:s2-3]
        tail = kks[s2-3:s2]
        second_list = kks[s2+1:e2-1].split(',')
        return base, first_list, body, tail, second_list

    def create_kks(
        self,
        head: str = "10",
        first_list = ["20", "30"],
        body: str = "KBF71AA",
        tail: str = "201",
        second_list = ["100", "200", "300"]):

        head_list = [head]
        head_list.extend(first_list)
        tail_list = [tail]
        tail_list.extend(second_list)

        kks = []
        for h in head_list:
            for t in tail_list:
                kks.append(str(h) + body + str(t))
        return kks

def get_stage_po(path):
    po = pd.read_excel(path)

    for i, row in po.iterrows():
        if "KKS Code" in row.to_list():
            kks_row = i
            kks_col = row.to_list().index("KKS Code")

    kks_index = []
    for i, kks in enumerate(po[po.columns[kks_col]].to_list()):
        if i >= kks_col and kks != '-' and not pd.isna(kks):
            kks_index.append(i)

    po.columns = po.loc[kks_row].to_list()
    po = po.loc[kks_index].reset_index(drop=True)

    po_dict = []
    for i, row in po.iterrows():
        dict_row = row.to_dict()
        dict_row["KKS Code"] = KKSTransformer()([row["KKS Code"]])[0]
        po_dict.append(dict_row)

    stage_po = pd.DataFrame(po_dict)
    return stage_po

def get_stage_cs(path):
    cs = pd.read_excel(path)
  
    for i, row in cs.iterrows():
        if "HNS Code" in row.to_list():
            hns_row = i
            hns_col = row.to_list().index("HNS Code")

    target_rows = []

    hns_index = []
    for i, hns in enumerate(cs[cs.columns[hns_col]].to_list()):
        if i >= hns_col and not pd.isna(hns):
            hns_index.append(i)

    cs.columns = cs.loc[hns_row].to_list()
    cs = cs.loc[hns_index].reset_index(drop=True)
    # ---------------------
    cs_dict = []

    last_non_null_kks = None
    last_non_null_project = None

    for i, row in cs.iterrows():

        if not pd.isna(row["KKS"]):
            parsed_kks = KKSTransformer()([row["KKS"]])
            last_non_null_kks = parsed_kks
        else:
            parsed_kks = last_non_null_kks

        if not pd.isna(row["Project #"]):
            last_non_null_project = row["Project #"]

        dict_row = {
            "number": row["№"],
            "eq_type": row["Equipment type**"],
            "eq": row["Equipment"],
            "conn_type": row.get("Connection type"),
            "qty": row["Quantity, pcs"],
            "project": last_non_null_project,
            "hns_code": row.get("HNS Code"),
            "kks": parsed_kks[0],
            "add_info": row.get("Additional information"),
            "weight_act": row.get('Weight of actuator (with bushing), kg'),
            "weight_trans": row.get('Weight of normalizing transducer, kg'),
            "total_weight": row.get('Total weight (with bushing), kg'),
            "contract_delivery_date": row.get('Contract delivery date'),
            "price_usd": row.get('Price, USD*'),
            "total_value_usd": row.get('Total value, USD*')
        }

        cs_dict.append(dict_row)
    
    stage_cs = pd.DataFrame(cs_dict)
    return stage_cs

def convert_date(dt, locale='ru'):
    months_ru = {
      1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля',
      5: 'мая', 6: 'июня', 7: 'июля', 8: 'августа',
      9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
    }

    months_en = {
      1: 'January', 2: 'February', 3: 'March', 4: 'March',
      5: 'May', 6: 'June', 7: 'July', 8: 'August',
      9: 'September', 10: 'October', 11: 'November', 12: 'December'
    }

    day, month, year = [int(v) for v in dt.split(".")]

    if locale == 'ru':
      formatted_date = f'«{day}» {months_ru[month]} {year} г.'
    elif locale == 'en':
      formatted_date = f'{months_en[month]} {day}, {year}'
    return formatted_date

def find_info_by_kks(kks_number, kks, stage_po, stage_cs):
  cs_rows = stage_cs[stage_cs['kks'].str.contains(kks)].to_dict(orient='records')
  po_rows = stage_po[stage_po['KKS Code'].str.contains(kks)].to_dict(orient='records')
  print(cs_rows)
  print(po_rows)
  assert len(po_rows) == 1, f"Problem with KKS=('{kks}')"

  ret = []
  for cs_row in cs_rows:
    ret.append([
        kks_number,
        "Сборка клапана с давальческим сырьем / Assembly of valve with the customer supplied equipment",
        po_rows[0]["Product type coding"],
        kks,
        cs_row["eq"],
        "шт. / pc.",
        cs_row["price_usd"],
        "1",
        cs_row["price_usd"],
        "1",
        cs_row["price_usd"],
        "0",
        "0"
    ])
  return ret

def start_text(sheet, constants):
    # Распаковка констант
    date = constants["date"]
    manuf_ru = constants["manuf_ru"]
    additional_agreement_no = constants["additional_agreement_no"]
    general_contract_ru = constants["general_contract_ru"]
    manuf_en = constants["manuf_en"]
    general_contract_en = constants["general_contract_en"]
    period_ru = constants["period_ru"]
    period_en = constants["period_en"]


    title = "Отчет об использовании давальческого сырья (материалов) / Report on the Use of Customer Supplied Equipment (materials)"
    city = "г. Санкт-Петербург / St. Petersburg"
    main_text = f"""Общество с ограниченной ответственностью «{manuf_ru}» (далее - Переработчик) для выполнения работ по Дополнительному соглашению {additional_agreement_no} к Рамочному договору {general_contract_ru}, составило настоящий отчет об использовании материалов (далее - Отчет), переданных обществом с ограниченной ответственностью «РЭС Инжиниринг» в лице Генерального директора Тихонова Сергея Сергеевича, действующего на основании Устава в редакции от «12» мая 2022 года, далее именуемое «Давалец» о том, что: /
    Dembla Valves Limited Company represented by J.N. Dembla (hereinafter referred to as the Processor), hereinafter referred to as the Supplying Customer to perform works under Supplementary Agreement {additional_agreement_no} to {general_contract_ru}, has drawn up this report on the use of equipment (hereinafter referred to as the "Report") handed over to RES Engineering Limited Liability Company represented by Sergey Tikhonov, General Director, acting on the basis of the Charter as amended on May 12, 2022, indicating that:"""
    first_point_text = "1.	Давалец передал, а Переработчик принял материалы для выполнения работ. / The Supplying Customer has handed over and the Processor has accepted the materials to perform the works."
    second_point_text = f"2.	В период {period_ru} переданные Заказчиком материалы использованы Подрядчиком при выполнении работ, а именно: / In the period {period_en}, the materials handed over by the Customer were used by the Contractor to perform the works, namely:"

    # --- 1. Title
    TITLE_ROW = 3
    sheet.cell(TITLE_ROW, 1).value = title
    sheet.cell(TITLE_ROW, 1).font = Font(bold=True, name="Times New Roman")
    sheet.cell(TITLE_ROW, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=13)

    # --- 2. City
    CITY_ROW = 5
    sheet.cell(CITY_ROW, 1).value = city
    sheet.cell(CITY_ROW, 1).font = Font(bold=False, name="Times New Roman")
    sheet.cell(CITY_ROW, 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=CITY_ROW, start_column=1, end_row=CITY_ROW, end_column=3)

    sheet.cell(CITY_ROW, 13).value = date
    sheet.cell(CITY_ROW, 13).font = Font(bold=False, name="Times New Roman")
    sheet.cell(CITY_ROW, 13).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=CITY_ROW, start_column=13, end_row=CITY_ROW, end_column=13)

    # --- 3. Main Text
    MAIN_TEXT_ROW = 7
    sheet.cell(MAIN_TEXT_ROW, 1).value = main_text
    sheet.cell(MAIN_TEXT_ROW, 1).font = Font(bold=False, name="Times New Roman")
    sheet.cell(MAIN_TEXT_ROW, 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=MAIN_TEXT_ROW, start_column=1, end_row=MAIN_TEXT_ROW, end_column=13)

    # --- 4. First Point
    FIRST_POINT_ROW = 8
    sheet.cell(FIRST_POINT_ROW, 1).value = first_point_text
    sheet.cell(FIRST_POINT_ROW, 1).font = Font(bold=False, name="Times New Roman")
    sheet.cell(FIRST_POINT_ROW, 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=FIRST_POINT_ROW, start_column=1, end_row=FIRST_POINT_ROW, end_column=13)

    # --- 5. Second Point
    SECOND_POINT_ROW = 9
    sheet.cell(SECOND_POINT_ROW, 1).value = second_point_text
    sheet.cell(SECOND_POINT_ROW, 1).font = Font(bold=False, name="Times New Roman")
    sheet.cell(SECOND_POINT_ROW, 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=SECOND_POINT_ROW, start_column=1, end_row=SECOND_POINT_ROW, end_column=13)

    sheet.row_dimensions[7].height = 70
    sheet.row_dimensions[9].height = 30

def end_text(sheet, end_idx, constants):

    amount_of_pieces = constants["amount_of_pieces"]
    kks = constants["kks"]
    kks = constants["kks"]
    product_type = constants["product_type"]
    invoice_1 = constants["invoice_1"]
    invoice_2 = constants["invoice_2"]
    invoice_3 = constants["invoice_3"]
    total_price = constants["total_price"]

    text_about_price = "Общая стоимость использованных материалов составила: / The total cost of the materials used amounted to:"
    total_price_as_words = "Write here"

    product_type_text = "3. Оборудование / Equipment"
    kks_text = f"в количестве {constants['amount_of_pieces']} штук, / in the amount of {constants['amount_of_pieces']} pieces,  KKS"

    after_kks_text = f"было собрано с использованием давальческого сырья, полученного в соответствии с инвойсом {invoice_1} для приводов и инвойсом {invoice_2} для нормирующих преобразователей и переходников, и передано заказчику в соответствии с инвойсом {invoice_3}. / was assembled using customer supplied equipment obtained according to invoice {invoice_1} for actuators and invoice {invoice_2} for normalizing transducers and bushings and handed over to the customer according to invoice {invoice_3}."
    fourth_point_text = "4. Настоящим подтверждаем, что неизрасходованных материалов, возвратных отходов нет, факта перерасхода материалов при выполнении работ не установлено.  / We hereby confirm that there are no unused materials, no returnables, and no overconsumption of the materials in the course of works was identified."
    fifth_point_text = "5. Настоящий Отчет составлен в 2 (двух) экземплярах, по одному для Подрядчика и Заказчика. / This Report is made in two (2) copies, each one for the Processor and the Supplying Customer."

    # --- 1.1 Text about price
    ROW_NUMER = end_idx
    sheet.cell(ROW_NUMER, 1).value = text_about_price
    sheet.cell(ROW_NUMER, 1).font = Font(bold=False, name="Times New Roman")
    sheet.cell(ROW_NUMER, 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=ROW_NUMER, start_column=1, end_row=ROW_NUMER, end_column=6)

    # --- 1.2 Total price
    ROW_NUMER = end_idx
    sheet.cell(ROW_NUMER, 7).value = "$" + str(total_price)
    sheet.cell(ROW_NUMER, 7).font = Font(bold=False, name="Times New Roman")
    sheet.cell(ROW_NUMER, 7).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

    # --- 1.3 Total price as words
    ROW_NUMER = end_idx
    sheet.cell(ROW_NUMER, 8).value = "Напиши цену словами"
    sheet.cell(ROW_NUMER, 8).font = Font(bold=False, name="Times New Roman")
    sheet.cell(ROW_NUMER, 8).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=ROW_NUMER, start_column=8, end_row=ROW_NUMER, end_column=13)

    # --- 2.1 Product type
    ROW_NUMER = end_idx + 2
    sheet.cell(ROW_NUMER, 1).value = product_type_text
    sheet.cell(ROW_NUMER, 1).font = Font(bold=False, name="Times New Roman")
    sheet.cell(ROW_NUMER, 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=ROW_NUMER, start_column=1, end_row=ROW_NUMER, end_column=2)

    # --- 2.2 Product type
    ROW_NUMER = end_idx + 2
    sheet.cell(ROW_NUMER, 3).value = product_type
    sheet.cell(ROW_NUMER, 3).font = Font(bold=False, name="Times New Roman")
    sheet.cell(ROW_NUMER, 3).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=ROW_NUMER, start_column=3, end_row=ROW_NUMER, end_column=13)

    # --- 3.1 KKS
    ROW_NUMER = end_idx + 3
    sheet.cell(ROW_NUMER, 1).value = kks_text
    sheet.cell(ROW_NUMER, 1).font = Font(bold=False, name="Times New Roman")
    sheet.cell(ROW_NUMER, 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=ROW_NUMER, start_column=1, end_row=ROW_NUMER, end_column=3)

    # --- 3.2 KKS
    ROW_NUMER = end_idx + 3
    sheet.cell(ROW_NUMER, 4).value = kks
    sheet.cell(ROW_NUMER, 4).font = Font(bold=False, name="Times New Roman")
    sheet.cell(ROW_NUMER, 4).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=ROW_NUMER, start_column=4, end_row=ROW_NUMER, end_column=13)

    # --- 4. AFTER KKS
    ROW_NUMER = end_idx + 4
    sheet.cell(ROW_NUMER, 1).value = after_kks_text
    sheet.cell(ROW_NUMER, 1).font = Font(bold=False, name="Times New Roman")
    sheet.cell(ROW_NUMER, 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=ROW_NUMER, start_column=1, end_row=ROW_NUMER, end_column=13)

    # --- 5. FOURTH POINT
    ROW_NUMER = end_idx + 5
    sheet.cell(ROW_NUMER, 1).value = fourth_point_text
    sheet.cell(ROW_NUMER, 1).font = Font(bold=False, name="Times New Roman")
    sheet.cell(ROW_NUMER, 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=ROW_NUMER, start_column=1, end_row=ROW_NUMER, end_column=13)

    # --- 6. FOURTH POINT
    ROW_NUMER = end_idx + 6
    sheet.cell(ROW_NUMER, 1).value = fifth_point_text
    sheet.cell(ROW_NUMER, 1).font = Font(bold=False, name="Times New Roman")
    sheet.cell(ROW_NUMER, 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=ROW_NUMER, start_column=1, end_row=ROW_NUMER, end_column=13)

    # --- 7. FROM DAVALETS
    ROW_NUMER = end_idx + 8
    sheet.cell(ROW_NUMER, 1).value = "От имени Давальца / On behalf of Supplying Customer"
    sheet.cell(ROW_NUMER, 1).font = Font(bold=False, name="Times New Roman")
    sheet.cell(ROW_NUMER, 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=ROW_NUMER, start_column=1, end_row=ROW_NUMER, end_column=4)

    # --- 8. FROM PERERABOTCHIK
    ROW_NUMER = end_idx + 14
    sheet.cell(ROW_NUMER, 1).value = "От имени Переработчика / On behalf of the Processor"
    sheet.cell(ROW_NUMER, 1).font = Font(bold=False, name="Times New Roman")
    sheet.cell(ROW_NUMER, 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    sheet.merge_cells(start_row=ROW_NUMER, start_column=1, end_row=ROW_NUMER, end_column=4)

def set_borders(sheet):
    # Define a border style with no borders
    no_border_side = Side(border_style=None)
    no_border = Border(left=no_border_side,
                    right=no_border_side,
                    top=no_border_side,
                    bottom=no_border_side)

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = None

def generate(stage_po, stage_cs, target_kks, head_constants, tail_constants):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Table"

    # --- СОЗДАЕМ ВЕРХНИЙ ТЕКСТ
    start_text(
        sheet=sheet,
        constants=head_constants
    )

    COLUMNS = [
        (1, 1, "No.", []),
        (2, 1, "Наименование вида работ / Type of work", []),
        (3, 1, "Тип продукции / Product type", []),
        (4, 1, "KKS код / KKS code", []),
        (5, 1, "Наименование материала / Name of material", []),
        (6, 1, "Единица измерения / Unit of measurement", []),
        (7, 1, "Стоимость за единицу измерения, USD / Cost per unit, USD", []),
        (8, 2, "Получено от Заказчика / Received from Customer", ["Кол-во / Quantity", "Сумма, USD/ Amount, USD"]),
        (10, 2, "Фактически использовано материалов / Materials actually used", ["Кол-во / Quantity", "Сумма, USD/ Amount, USD"]),
        (12, 2, "Остаток неиспользованных материалов / Balance of unused materials", ["Кол-во / Quantity", "Сумма, USD/ Amount, USD"])
    ]

    # --- Создаем голову
    HEAD_INDEX = 11

    for column in COLUMNS:
        i, width, col_name, sub_cols = column

        sheet.cell(HEAD_INDEX, i).value = col_name

        if width > 1:
            sheet.merge_cells(start_row=HEAD_INDEX, start_column=i, end_row=HEAD_INDEX, end_column=i+width-1)

        if sub_cols is not None:
            if len(sub_cols) != 0: # Если есть дополнительные колонки
                for shift, sub_col in enumerate(sub_cols):
                    sheet.cell(HEAD_INDEX+1, i + shift).value = sub_col
        else: # Если нет дополнительных колонок
            sheet.merge_cells(start_row=HEAD_INDEX, start_column=i, end_row=HEAD_INDEX+1, end_column=i)

    # --- FILL INFO
    prices = []
    row_idx = HEAD_INDEX + 2
    for kks_number, kks in enumerate(target_kks):
        rows_group = find_info_by_kks(kks_number+1, kks, stage_po, stage_cs)

        if len(rows_group) == 0: continue
    
        for row in rows_group:
            for i, elem in enumerate(row):
                sheet.cell(row_idx, i+1).value = elem
            row_idx += 1
            prices.append(row[8])
        for col_number in [1, 2, 3, 4]:
            sheet.merge_cells(
                start_row=row_idx-len(rows_group),
                start_column=col_number,
                end_row=row_idx-1,
                end_column=col_number
            )

    # --- END TEXT
    tail_constants["total_price"] = sum(prices)
    end_text(
        sheet,
        row_idx+1,
        constants=tail_constants
    )

    # --- STYLE
    RANGES = [
        { # HEAD
            "coords": {"start_row": HEAD_INDEX, "start_column": 1, "end_row": HEAD_INDEX+1, "end_column": 14},
            "font": Font(bold=True, name='Times New Roman'),
            "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True)
        },
        { # BODY
            "coords": {"start_row": HEAD_INDEX+2, "start_column": 1, "end_row": row_idx, "end_column": 14},
            "font": Font(bold=False, name="Times New Roman"),
            "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True)
        }
    ]

    for style_range in RANGES:
        for r_idx in range(style_range["coords"]["start_row"], style_range["coords"]["end_row"] + 1):
            for c_idx in range(style_range["coords"]["start_column"], style_range["coords"]["end_column"] + 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                cell.alignment = style_range["alignment"]
                cell.font = style_range["font"]

    # --- BORDERS
    set_borders(sheet)

    # --- COL WIDTH
    cols_width = [10, 25, 25, 25, 25, 25, 25, 10, 25, 10, 25, 10, 25]
    for i, col_width in enumerate(cols_width):
        column_letter = get_column_letter(i + 1)
        sheet.column_dimensions[column_letter].width = col_width

    # --- HEAD COLS WIDTH
    sheet.row_dimensions[HEAD_INDEX].height = 50
    sheet.row_dimensions[HEAD_INDEX + 1].height = 50
    workbook.save("joined.xlsx")

def download_file(url, save_path):
    # url = "https://v5.airtableusercontent.com/v3/u/48/48/1764936000000/XA0lfHHpFysZTPcsspUfnw/PjFavBdktf_bbl85nsSU-q_ezlPaB_52kVvxeEwSK4Gft4tnyisAwnk1Vl07wBU3MJH0FYdZGlCbU55PFbBNZFZWFJX-s1Ux8meZm_Mr0I4CfJRaGw8SOBwpOSTdaVs-L3VXEebV-agJe2DZg1kwLTcCJweJUg9epccEq2AKTRg55r8jbx4VWMjyb6O4DxOy/sheu7wZolfmwLMSM5YGoj6-VzWFnFPjlnVNymM901gk"

    print("FETCHING", url)
    response = requests.get(url, stream=True)
    response.raise_for_status()  # Проверяем успешность запроса
    #save_path = "lol.xlsx"
    print("STATUS", response.status_code)
    # Сохраняем файл
    with open(save_path, 'wb') as file:
        for chunk in response.iter_content(chunk_size=8192):
            print('chunk')
            file.write(chunk)
@task
def fetch_and_generate():
    autogenerate_table = get_table("Portal", "Portal 2.0.3", "Autogenerated docs")

    rows = autogenerate_table.all(
        formula="NOT({Script check})",
        cell_format="string",
        time_zone="Europe/Moscow",  # Укажите ваш часовой пояс
        user_locale="ru"  # Укажите вашу локаль
    )

    # === Идем по строкам из AT ===
    for row in rows:

        # === id записи ===
        record_id = row["id"]
        print("------>>> RECORD ID", record_id)

        # === KKS которые запросил прользователь ===
        target_kks = row["fields"]["KKS"].split(", ")
        print("------>>> TARGET KKS", target_kks)

        # === Получаем ссылки на файлы ===
        po_url = rows[0]["fields"]["PO from Excel"]
        po_url = po_url[po_url.find("https"):-1]

        cs_url = rows[0]["fields"]["CS from Excel"]
        cs_url = cs_url[cs_url.find("https"):-1]

        # === Скачиваем файлы ===
        download_file(po_url, "po.xlsx")
        download_file(cs_url, "cs.xlsx")
        print("------>>> DOWNLOADED")

        # === Парсим файлы ===
        stage_po = get_stage_po("po.xlsx")
        stage_cs = get_stage_cs("cs.xlsx")
        print("------>>> STAGED")

        print("STAGE PO COLUMNS", stage_po.columns)
        print("STAGE CS COLUMNS", stage_cs.columns)

        # === Конвертация дат ===
        ru_dt1 = convert_date(row["fields"]["Start date"], locale="ru")
        ru_dt2 = convert_date(row["fields"]["End date"], locale="ru")

        en_dt1 = convert_date(row["fields"]["Start date"], locale="en")
        en_dt2 = convert_date(row["fields"]["End date"], locale="en")

        manuf_translate_ru = {
            "Dembla": "«Dembla Valves Limited Company» в лице J.N. Dembla"
        }
        manuf_translate_en = {
            "Dembla": "Dembla Valves Limited Company represented by J.N. Dembla"
        }

        # === Формируем константы для генерации документа ===
        head_constants = {
            "date": "10.08.2025",
            "manuf_ru": manuf_translate_ru[row["fields"]["Manufacturer"]],
            "additional_agreement_no": row["fields"]["Additional agreement number"],
            "general_contract_ru": row["fields"]["GC rus"],
            "manuf_en": manuf_translate_en[row["fields"]["Manufacturer"]],
            "general_contract_en": row["fields"]["GC eng"],
            "period_ru": f"с {ru_dt1} по {ru_dt2}",
            "period_en": f"{en_dt1}. and {en_dt2}"
        }

        tail_constants = {
            "amount_of_pieces": len(target_kks) * 3,
            "kks": ", ".join(target_kks),
            "product_type": ", ".join(stage_po["Product type coding"].unique()),
            "invoice_1": row["fields"]["Actuator's invoice"],
            "invoice_2": row["fields"]["Converters and adapters invoice"],
            "invoice_3": row["fields"]["Consignment notes' invoice"]
        }

        generate(
            stage_po=stage_po,
            stage_cs=stage_cs,
            target_kks=target_kks,
            head_constants=head_constants,
            tail_constants=tail_constants
        )

        service = create_service()
        to_email = row["fields"]["Creator's email"], #"letyagin.a@res-e.ru"
        subject = "test_subject"
        body = "test_body"
        excel_file_path = "joined.xlsx"
        send_email_with_excel(service, to_email, subject, body, excel_file_path)

        autogenerate_table.update(record_id, {"Script check": True})

with DAG(
    dag_id="davalcheskoe",
    default_args=default_args,
    start_date=START_DATE,
    schedule_interval=CRON_EXP,
    catchup=False
) as dags:
    
    fetch_and_generate()
